"""
    A discovery and collection tool which pulls information from switches and
    provides useful data points and recommendations on which ports
    to apply or exclude 802.1x NAC configurations. Information is exported to an
    Excel workbook.

    Currently Recommendations based on based on LLDP Vendorlookup, RemoteCapability,
    Description keywords, and multiple macs present on port.
"""

from nornir import InitNornir
from nornir.core.exceptions import NornirExecutionError
from nornir.core.exceptions import NornirSubTaskError
from nornir.plugins.tasks import networking
from nornir.plugins.tasks.networking import napalm_get
from nornir.plugins.functions.text import print_result
from mac_vendor_lookup import MacLookup
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font
import re

def create_workbook():
    """
    Create an Excel workbook to store values retrieved from switches
    """

    wb = openpyxl.Workbook()
    groupname = "Grouping1" #TODO: Replace with function that takes list of location codes.
    wb_name = "NACFACTS -" + groupname + ".xlsx"


    #Create sheets and column headers
    facts_ws = wb.create_sheet("Facts")
    facts_ws.append(['Switch Hostname','Vendor','Model','OS Version','Serial Number','Uptime'])

    interfaces_ws = wb.create_sheet("Interfaces")
    interfaces_ws.append(['Switch', 'Interface name', 'Description', 'Admin Status', 'Oper Status', 'Speed'])

    mactablevendor_ws = wb.create_sheet("Mac Table Vendors")
    mactablevendor_ws.append(['Switch', 'Interface', 'MACaddr', 'Vendor OUI'])

    lldpneighbor_ws = wb.create_sheet("LLDP Neighbors")
    lldpneighbor_ws.append(['Local Switch', 'Local Port', 'Remote System ID', 'Remote System Name', 'Remote System Description', 'Remote Port ID', 'Remote Port Description', 'Remote Capability', 'Remote Vendor'])

    multimacports_ws = wb.create_sheet("Multi Mac Ports")
    multimacports_ws.append(['Switch', 'Interface', 'Count', 'Vendor MACs'])

    portexclusions_ws = wb.create_sheet("Port Exclusion Recommendations")
    portexclusions_ws.append(['Switch', 'Interface', 'Reason', 'port description'])

    """
    Initialize Nornir settings and set the right inventory targets and filters
    """
    nr = InitNornir(config_file="config.yaml", core={"raise_on_error": False})
    #accessHosts = nr.filter(hostname='10.83.8.163')
    accessHosts = nr.filter(site='herndon-dev')
    #accessHosts = nr.filter(type='network_device')
    #accessHosts = nr.filter(role='FIAB')
    #accessHosts = nr.filter(site='home')


    #Initialize nested dictionary for tracking recomended ports and reasoning to exclude from NAC.
    portexclusions = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    #portexclusions = {'SwitchName': {'Gi1/1:' {'description': 'trunk to mdf', macvendor: 'VMWare', Reasoning: ['Port is trunk', 'multimacs'] }}

    """
    Get Mac Table results from all inventory targets using nornir napalm
    task=get_mac_table, lookup Vendor MACOUI and store in mactablevendor_ws
    """
    #mac_results = accessHosts.run(task=get_mac_table_napalm, name="Get MAC Table")
    #print('Print hosts before .run\n', nr.inventory)
    #print(len(accessHosts))
    print('Collecting information from the following Nornir inventory hosts:', accessHosts.inventory.hosts.keys())
    #try:
    mac_results = accessHosts.run(task=napalm_get, getters=['mac_address_table'])
    #except Exception as e:
        #print('ERROR!!!\n\n', e)

    macQ = MacLookup()
    #loop through each switch in the Nornir Task Result object
    for host, task_results in mac_results.items():
        print("Start processing Host - Mac_Results:", str(host), '\n')
        #Check for switches that failed and pass. Nornir automatically removes failed devices from future tasks.
        #TODO: Log or print failures for further review.
        if task_results.failed:
            print(' ######', '\n!Failed Host in task_results:>', host, 'will be removed from future tasks!\n', '######', '\n')
            continue
        #Store the actual serialized mac table dict for the current switch
        mac_results_host =  task_results[0].result

        vendor_mactable = defaultdict(list)
        interfaces = defaultdict(list)
        multimacinterfaces = defaultdict(dict)
        #Loop through each Host's MAC Table and Create dictionary of interfaces and the vendor MACs that are attached.
        for entry in mac_results_host['mac_address_table']:
            if not entry['interface']:  #skip mac address not assigned to interfaces
                continue
            try:
                vendor = macQ.lookup(entry['mac'])
            except:
                vendor = "Unknown"

            interface_value = entry['interface']
            vendor_value = vendor
            mac_value = entry['mac']

            #Store relevant values for worksheet row and append to sheet.
            line = [host, interface_value, mac_value, vendor_value]
            mactablevendor_ws.append(line)
            #Append Vendor lookup results to vendor_mactable so we can use them for port exclusion recommendations.
            vendor_mactable[entry['interface']].append(vendor_value)

        #build dictionary of interfaces containing lists of vendors and identify ports with multiple MACs.
        for iface, value in vendor_mactable.items():
            #print(iface, value)
            if len(value) > 1:
                #print(iface, '>', value)
                interfaces[iface].extend(value)

                line = [host, iface, len(interfaces[iface]), str(interfaces[iface])]
                #print(line)
                multimacports_ws.append(line)
                #Append to portexlcusions dictionary
                portexclusions[host][iface]['reason'].append('multimac')
        #print('vendor mactable\n\n', vendor_mactable)
        #print('interfact dict \n\n', interfaces)
        print("End Processing Host - Mac_Results: " + str(host) + "\n")
    #exit()
    """
    Get Facts  from all inventory targets using nornir napalm
    task=get_facts and output results to facts_ws
    """
    facts = accessHosts.run(task=napalm_get, getters=["facts"])
    #print('Check nornir inventory after potential failure > accesshosts keys:>', accessHosts.inventory.hosts.keys())

    #Loop through each host's task results and store values to append lines to facts_ws
    #AggregatedResult (napalm_get): {'C9300-48-UXM-1': MultiResult: [Result: "napalm_get"], 'C9500-16X': MultiResult: [Result: "napalm_get"]}
    for host, task_results in facts.items():
        print("Start processing Host - Get Facts:", str(host), '\n')
        facts_result =  task_results[0].result

        vendor_result = facts_result['facts']['vendor']
        model_result = facts_result['facts']['model']
        version_result = facts_result['facts']['os_version']
        serial_result = facts_result['facts']['serial_number']
        uptime_result = facts_result['facts']['uptime']

        line = [host, vendor_result, model_result, version_result, serial_result, uptime_result]

        facts_ws.append(line)
        print("End Processing Host - Get Facts: " + str(host) + "\n")

    """
    get_lldp_neighbors for all invetnory targets using nornir napalm task=get_lldp_neighbors,
    perform mac vendoroui lookup on chassisid and output all results to lldpneighbors_ws
    """

    lldpneighbors = accessHosts.run(task=napalm_get, getters=["lldp_neighbors_detail"])

    for host, task_results in lldpneighbors.items():
        print("Start processing Host - Get LLDP Neighbors:", str(host), '\n')
        #Store results from Nornir aggregated result object
        lldp_result =  task_results[0].result
        #store actual result dicitonary from the Nornir result object.
        lldp_detail = lldp_result['lldp_neighbors_detail']

        for interface in lldp_detail:
            #print(lldp_detail)

            remotesysid = lldp_detail[interface][0]['remote_chassis_id']
            remotesysname = lldp_detail[interface][0]['remote_system_name']
            remotesysdescription = lldp_detail[interface][0]['remote_system_description']
            remoteportid = lldp_detail[interface][0]['remote_port']
            remoteportdesc = lldp_detail[interface][0]['remote_port_description']
            remotecapability = lldp_detail[interface][0]['remote_system_capab']
            try:
                remotevendor = macQ.lookup(remotesystemid)
            except:
                remotevendor = "Unknown"

            line = [host, interface, remotesysid, remotesysname, remotesysdescription, remoteportid, remoteportdesc, str(remotecapability), remotevendor]
            #print(line)
            lldpneighbor_ws.append(line)

            if ('router' in remotecapability) or ('bridge' in remotecapability):
                #TODO: generalize for all interfaces and move to function
                if re.search('TenGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Te' + digits.group()
                elif re.search('TwoGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Tw' + digits.group()
                elif re.search('^Gigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Gi' + digits.group()

                portexclusions[host][interface]['reason'].append('LLDP Neighbor' + str(remotecapability))
                #print(host, interface, remotecapability)

        print("End Processing Host - Get LLDP Neighors: " + str(host) + "\n")

    """
    Get Interfaces, check descriptions for keywords and append to port exclusions.
    """
    getinterfaces = accessHosts.run(task=napalm_get, getters=["interfaces"])

    for host, task_results in getinterfaces.items():
        print("Start processing Host - Get Interfaces:", str(host), '\n')

        interfaces_result = task_results[0].result
        interfaces_result = interfaces_result['interfaces']
        for interface in interfaces_result:
            #if 'C9500-16X' in host:
                #print(interface_id)
            interface_id = interface
            adminstatus = interfaces_result[interface]['is_enabled']
            operstatus = interfaces_result[interface]['is_up']
            description = interfaces_result[interface]['description']
            speed = interfaces_result[interface]['speed']

            line = [host, interface_id, description, adminstatus, operstatus, speed]
            #print(line)

            interfaces_ws.append(line)

            #Check for Exclusion keywords and add interfaces to portexclusion dictionary then append to portexlusion_ws.
            #TODO: Replace search literals with variable at top for quicker modification.
            keyword = re.search('(ASR|ENCS|UPLINK|CIRCUIT|ISP|SWITCH|TRUNK|ESXI|VMWARE)', str(description), re.IGNORECASE)
            if keyword:
                #Normalize Interface names because different napalm getters return full interfaces name and some return shortened names which result in multiple dictionary keys being created.
                #TODO: generalize for all interfaces and move to function
                if re.search('TenGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Te' + digits.group()
                elif re.search('TwoGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Tw' + digits.group()
                elif re.search('^Gigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Gi' + digits.group()

                reasondescript = 'Description contains: ' + keyword.group()
                portexclusions[host][interface]['reason'].append(reasondescript)
                portexclusions[host][interface]['description']= str(description)

        print("End processing Host - Get Interfaces:", str(host), '\n')

    """
    Export all entries from portexlusions dictionary to portexclusions_ws
    so that port exclusion recommendations show up in the workbook.
    """
    for host, value in portexclusions.items():
        print("Start processing Host - Port Exclusions:", str(host), '\n')
        #print(host)
        for interface in portexclusions[host]:
            #print(host, interface, portexclusions[host][interface]['reason'], portexclusions[host][interface]['description'])
            line = [host, interface, str(portexclusions[host][interface]['reason']), str(portexclusions[host][interface]['description'])]
            portexclusions_ws.append(line)
        print("End processing Host - Port Exclusions:", str(host), '\n')

    """
    Get VLANs... in Napalm-automation:develop train, needed for identifying switchport mode trunk.
    """
    #vlans = accessHosts.run(task=napalm_get, getters=['vlan'], name="Get VLANs")
    #print_result(vlans)

    if nr.data.failed_hosts:
        print("The following switches failed during a task and were not added to the workbook:", nr.data.failed_hosts)
    #TODO: save failed switches to worksheet

    wb.remove(wb["Sheet"])
    #catch potential save errors on workbook.
    try:
        wb.save(wb_name)
        print("\nWorkbook Created")
    except Exception as e:
        print("\n######", e , "######\nFailed to Save workbook, please close it if open and ensure you have access to save location")

"""Password Handler to avoid storing credentials in the clear inside inventory files."""
def nornir_set_creds(norn, username=None, password=None):
    if not username:
        username = input("Enter username: ")
    if not password:
        password = getpass()

    for host_obj in norn.inventory.hosts.values():
        host_obj.username = username
        host_obj.password = password
"""
Run the main function to pull device information and create the workbook.
"""
create_workbook()
