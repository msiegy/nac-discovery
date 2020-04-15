from nornir import InitNornir
from nornir.plugins.tasks import networking
from nornir.plugins.functions.text import print_result
from mac_vendor_lookup import MacLookup
from collections import defaultdict
import openpyxl

def create_workbook():
    """
    Create an Excel workbook to store values retrieved from switches
    """

    wb = openpyxl.Workbook()
    switchhostname = "switch"
    wb_name = "NACFACTS -" + switchhostname + ".xlsx"

    #Create sheets and column headers
    facts_ws = wb.create_sheet("Facts")
    facts_ws.append(['Hostname','Vendor','Model','OS Version','Serial Number','Uptime'])

    portexclusions_ws = wb.create_sheet("Port Exclusion Recommendations")
    portexclusions_ws.append(['Interface', 'Description', 'Switchport Mode', 'MAC OUI Vendor', 'MACaddr', 'Reason'])

    lldpneighbor_ws = wb.create_sheet("LLDP Neighbors")
    lldpneighbor_ws.append(['Local Hostname', 'Local Port', 'Remote Hostname', 'Remote Port Description' 'Remote Vendor'])

    nr = InitNornir(config_file="config.yaml")
    accessHosts = nr.filter(hostname='10.83.8.163')

    lldp_results = accessHosts.run(task=get_neighbors, name="Find LLDP Neighbors")

    mac_results = accessHosts.run(task=get_mac_table, name="Get MAC Info and Vendors")

    print(mac_results, '\n\n')
    print(mac_results['C9300-48-UXM-1'][0])


    #portexclusions = mac_results['C9300-48-UXM-1'][0][1]
    #print ('EXCLUSIONS', portexlusions)
    wb.remove(wb["Sheet"])
    wb.save(wb_name)

def get_mac_table(task):
    r = task.run(networking.napalm_get, getters=['mac_address_table'], name="Get MAC Table")
    macQ = MacLookup()

    print('-'*10, task.host, '-', task.host.hostname, '- MultiMAC Interfaces','-'*10)
    #transform MAC address table into a new dictionary of interfaces with a list of associated MACs per interface. Only Keep multi mac interfaces
    #task.host['MultiMacports'] = multimacports(r.result)
    #Lookup Vendor OUI from MAC
    vendor_mactable = defaultdict(list)
    interfaces = defaultdict(list)

    #Create dictionary of interfaces and the vendor MACs that are attached.
    for key in r.result['mac_address_table']:
        if not key['interface']:
            continue

        try:
            vendor = macQ.lookup(key['mac'])
        except:
            vendor = "Unknown - " + key['mac']

        #print(key['mac'], 'EQUALS > ', vendor, 'ON INTERFACE > ', key['interface'])
        vendor_mactable[key['interface']].append(vendor)

    #From VendorMacTable dictionary, create new list of interfaces that have multiple MACs.
    #Add to Exclusion List
    for iface, value in vendor_mactable.items():
        #print(iface, value)
        if len(value) > 1:
            #print(iface, '>', value)
            interfaces[iface].extend(value)

    print('----- Resolved Vendor MAC Table ----- \n', vendor_mactable)
    print('----- Multi Host Interfaces ----- \n', interfaces)

    task.host['vendormactable'] = vendor_mactable
    task.host['multimacinterfaces'] = interfaces
    #return task.host['exlcusions'][vendor_mactable, interfaces]
    info = {'multimacports': dict(interfaces), 'vendormactable': dict(vendor_mactable)}
    #info = {'multimacports': [0,1,2,3,4,5]}
    return info

def get_neighbors(task):
    r = task.run(networking.napalm_get, getters=['lldp_neighbors_detail'], name="Get LLDP Neighbor Details")
    # save our values in to the Key 'neighbors'
    task.host['neighbors'] = r.result
    #loop through the neighbors
    #print(task.host['neighbors']['lldp_neighbors_detail'])
    print('-'*10, task.host, '-', task.host.hostname, '- LLDP Neighbors', '-'*10)
    for neighbor in task.host['neighbors']['lldp_neighbors_detail']:
        remotemac = task.host['neighbors']['lldp_neighbors_detail'][neighbor][0]['remote_chassis_id']
        remotesystem = task.host['neighbors']['lldp_neighbors_detail'][neighbor][0]['remote_system_name']
        remotedescription = task.host['neighbors']['lldp_neighbors_detail'][neighbor][0]['remote_port_description']
        try:
            remotemacOUI = macQ.lookup(remotemac)
        except Exception as e:
            remotemacOUI = 'Vendor Unknown'
        localport = neighbor
        #print("Local Port:", localport, "Remote System:", remotesystem, "Remote Port Desc", remotedescription, "LLDP Neighbor MAC:", remotemac, "VENDOR:", remotemacOUI)

        line = [task.host, localport, remotesystem, remotedescription, remotemacOUI]


def get_interfaces(task):
    r = task.run(networking.napalm_get, getters=['interfaces'], name="Get Interfaces")
    # save our values in to the Key 'neighbors'
    task.host['interfaces'] = r.result
    #loop through the neighbors
    print(task.host['interfaces'])
    #print('-'*10, task.host, '-', task.host.hostname, '-'*10)

create_workbook()
