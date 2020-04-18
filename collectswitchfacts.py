"""
    A discovery and collection tool which pulls information from switches and
    provides useful data points and recommendations on which ports
    to apply or exclude 802.1x NAC configurations. Information is exported to an
    Excel workbook.

    Currently Recommendations based on based on LLDP Vendorlookup, RemoteCapability,
    Description keywords, and multiple macs present on port.
"""

from nornir import InitNornir
from nornir.core.exceptions import NornirExecutionError
from nornir.core.exceptions import NornirSubTaskError
from nornir.plugins.tasks import networking
from nornir.plugins.tasks.networking import napalm_get
from nornir.plugins.functions.text import print_result
from mac_vendor_lookup import MacLookup
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font
import re

def create_workbook():
    """
    Create an Excel workbook to store values retrieved from switches
    """

    wb = openpyxl.Workbook()
    groupname = "failedGrouping" #TODO: Replace with function that takes list of location codes.
    wb_name = "NACFACTS -" + groupname + ".xlsx"


    #Create sheets and column headers
    facts_ws = wb.create_sheet("Facts")
    facts_ws.append(['Switch Hostname','Vendor','Model','OS Version','Serial Number','Uptime'])

    interfaces_ws = wb.create_sheet("Interfaces")
    interfaces_ws.append(['Switch', 'Interface name', 'Description', 'Admin Status', 'Oper Status', 'Speed'])

    mactablevendor_ws = wb.create_sheet("Mac Table Vendors")
    mactablevendor_ws.append(['Switch', 'Interface', 'MACaddr', 'Vendor OUI'])

    lldpneighbor_ws = wb.create_sheet("LLDP Neighbors")
    lldpneighbor_ws.append(['Local Switch', 'Local Port', 'Remote System ID', 'Remote System Name', 'Remote System Description', 'Remote Port ID', 'Remote Port Description', 'Remote Capability', 'Remote Vendor'])

    multimacports_ws = wb.create_sheet("Multi Mac Ports")
    multimacports_ws.append(['Switch', 'Interface', 'Count', 'Vendor MACs'])

    portexclusions_ws = wb.create_sheet("Port Exclusion Recommendations")
    portexclusions_ws.append(['Switch', 'Interface', 'Reason', 'port description'])

    """
    Initialize Nornir settings and set the right inventory targets and filters
    """
    nr = InitNornir(config_file="config.yaml", core={"raise_on_error": False})
    #accessHosts = nr.filter(hostname='10.83.8.163')
    #accessHosts = nr.filter(site='herndon-dev')
    #accessHosts = nr.filter(type='network_device')
    accessHosts = nr.filter(role='FIAB')

    #Initialize nested dictionary for tracking recomended ports and reasoning to exclude from NAC.
    portexclusions = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    #portexclusions = {'SwitchName': {'Gi1/1:' {'description': 'trunk to mdf', macvendor: 'VMWare', Reasoning: ['Port is trunk', 'multimacs'] }}

    """
    Get Mac Table results from all inventory targets using nornir napalm
    task=get_mac_table_napalm, lookup Vendor MACOUI and store in mactablevendor_ws
    """
    #mac_results = accessHosts.run(task=get_mac_table_napalm, name="Get MAC Table")
    #print('Print hosts before .run\n', nr.inventory)
    #print(len(accessHosts))
    print('nornir inventory accesshosts keys:>', accessHosts.inventory.hosts.keys())
    #try:
    mac_results = accessHosts.run(task=napalm_get, getters=['mac_address_table'])
    print('print lenght of macresults ', len(mac_results))
    print('MAC RESULTS>', mac_results)
    #except Exception as e:
#        print('ERROR!!!\n\n', e)
#        print(e)

    #print('MAC RESULTS OUTPUT >>\n', mac_results)
    #print_result(mac_results)
    macQ = MacLookup()
    #print('Pre Loop, print results from run task:\n', mac_results)
    #print('Pre Loop, print_results\n', print_result(mac_results))
    #loop through each switch in the Nornir Task Result object
    for host, task_results in mac_results.items():
        print("Start processing Host - Mac_Results:", str(host), '\n')
        #Check for switches that failed and pass... TODO: Log failures for further review.
        if task_results.failed:
            print('Failed Host in task_results:>', host)
            continue
        #Return the actual serialized mac table dict for the current switch
        mac_results_host =  task_results[0].result

        vendor_mactable = defaultdict(list)
        interfaces = defaultdict(list)
        multimacinterfaces = defaultdict(dict)
        #print(mac_results_host)
        #print('Enter loop: for host, task results in mac_Results.items() loop... Print mac_results_host\n', mac_results_host)
        #Loop through each Host's MAC Table and Create dictionary of interfaces and the vendor MACs that are attached.

        #print('Entered MAC Results table loop. Print mac_results_host object\n', mac_results_host)
        #Loop through the list of Mac Table Entries and create a dictionary of interfaces and Lookup the vendor MACs that are attached.
        #print('Print current Hosttable loop object\n', hosttable)
        #print(mac_results_host)
        for entry in mac_results_host['mac_address_table']:
            #print(entry)
            if not entry['interface']:  #skip mac address not assigned to interfaces
                continue
            try:
                vendor = macQ.lookup(entry['mac'])
            except:
                vendor = "Unknown"

            #print(entry['mac'], 'EQUALS > ', vendor, 'ON INTERFACE > ', entry['interface'])
            #vendor_mactable[entry['interface']].append(vendor)
            interface_value = entry['interface']
            vendor_value = vendor
            mac_value = entry['mac']

            #Store relevant values for worksheet row
            line = [host, interface_value, mac_value, vendor_value]
            #print(line)
            mactablevendor_ws.append(line)
            #
            vendor_mactable[entry['interface']].append(vendor_value)

        #build dictionary of interfaces containing lists of vendors and identify ports with multiple MACs.
        for iface, value in vendor_mactable.items():
            #print(iface, value)
            if len(value) > 1:
                #print(iface, '>', value)
                interfaces[iface].extend(value)

                line = [host, iface, len(interfaces[iface]), str(interfaces[iface])]
                #print(line)
                multimacports_ws.append(line)
                #Append to portexlcusions dictionary
                portexclusions[host][iface]['reason'].append('multimac')
        #print('vendor mactable\n\n', vendor_mactable)
        #print('interfact dict \n\n', interfaces)
        print("End Processing Host - Mac_Results: " + str(host) + "\n")
    #exit()
    """
    Get Facts  from all inventory targets using nornir napalm
    task=get_facts and output results to facts_ws
    """
    facts = accessHosts.run(task=get_facts, name="Get Facts")
    print('nornir inventory accesshosts keys:>', accessHosts.inventory.hosts.keys())
    print('print lenght of FACTS ', len(facts))
    print('FACTS>', facts)

    #print(facts.items())

    """
    Loop through this below.... Each Device has amultiresult dictionary with list of dictionarys.
    dict_items([('C9300-DNA-A-24', MultiResult: [Result: "Get Facts", Result: "Get facts"]), ('C9300-48-HXEdge', MultiResult: [Result: "Get Facts", Result: "Get facts"]), ('C9300-48-UXM-1', MultiResult: [Result: "Get Facts", Result: "Get facts"]), ('C9500-16X', MultiResult: [Result: "Get Facts", Result: "Get facts"])])

    C9300-DNA-A-24
    C9300-48-HXEdge
    C9300-48-UXM-1
    C9500-16X

    """
    """
    for dev in facts:
        print(dev)
        #model = facts[dev].result['facts']['model']
    """
    #Loop through each host's task results and store values to append lines to facts_ws
    for host, task_results in facts.items():
        print("Start processing Host - Get Facts:", str(host), '\n')
        facts_result =  task_results[1].result

        vendor_result = facts_result['facts']['vendor']
        model_result = facts_result['facts']['model']
        version_result = facts_result['facts']['os_version']
        serial_result = facts_result['facts']['serial_number']
        uptime_result = facts_result['facts']['uptime']

        line = [host, vendor_result, model_result, version_result, serial_result, uptime_result]

        facts_ws.append(line)
        print("End Processing Host - Get Facts: " + str(host) + "\n")

    """
    get_lldp_neighbors for all invetnory targets using nornir napalm task=get_lldp_neighbors,
    perform mac vendoroui lookup on chassisid and output all results to lldpneighbors_ws
    """

    lldpneighbors = accessHosts.run(task=get_lldp_neighbors, name="Get LLDP Neighbors")

    for host, task_results in lldpneighbors.items():
        print("Start processing Host - Get LLDP Neighbors:", str(host), '\n')
        #Store results from Nornir aggregated result object
        lldp_result =  task_results[1].result
        #store actual result dicitonary from the Nornir result object.
        lldp_detail = lldp_result['lldp_neighbors_detail']

        for interface in lldp_detail:
            #print(lldp_detail)

            remotesysid = lldp_detail[interface][0]['remote_chassis_id']
            remotesysname = lldp_detail[interface][0]['remote_system_name']
            remotesysdescription = lldp_detail[interface][0]['remote_system_description']
            remoteportid = lldp_detail[interface][0]['remote_port']
            remoteportdesc = lldp_detail[interface][0]['remote_port_description']
            remotecapability = lldp_detail[interface][0]['remote_system_capab']
            try:
                remotevendor = macQ.lookup(remotesystemid)
            except:
                remotevendor = "Unknown"

            line = [host, interface, remotesysid, remotesysname, remotesysdescription, remoteportid, remoteportdesc, str(remotecapability), remotevendor]
            #print(line)
            lldpneighbor_ws.append(line)

            if 'router' or 'bridge' in remotecapability:
                portexclusions[host][interface]['reason'].append('LLDP Neighbor' + str(remotecapability) )

        print("End Processing Host - Get LLDP Neighors: " + str(host) + "\n")

    """
    Get Interfaces... May not be necessary, has description, Up/Down info, but need VLANs
    """
    getinterfaces = accessHosts.run(task=get_interfaces, name="Get Interfaces")

    for host, task_results in getinterfaces.items():
        print("Start processing Host - Get Interfaces:", str(host), '\n')

        interfaces_result = task_results[1].result
        interfaces_result = interfaces_result['interfaces']
        for interface in interfaces_result:
            #if 'C9500-16X' in host:
                #print(interface_id)
            interface_id = interface
            adminstatus = interfaces_result[interface]['is_enabled']
            operstatus = interfaces_result[interface]['is_up']
            description = interfaces_result[interface]['description']
            speed = interfaces_result[interface]['speed']

            line = [host, interface_id, description, adminstatus, operstatus, speed]
            #print(line)

            interfaces_ws.append(line)

            #Check for Exlcusion keywords and add switch and interfaces to portexclusion dictionary, append to portexlusion_ws later.
            keyword = re.search('(ASR|ENCS|UPLINK|CIRCUIT|ISP|SWITCH|TRUNK|ESXI|VMWARE)', str(description), re.IGNORECASE)
            if keyword:
                #Different napalm getters return full interfaces name and some return shortened names which result in multiple dictionary keys being created.
                #Improve logic handling here for any interface type, move to function...
                if "TwoGigabit" in interface:
                    interface = "Tw" + interface[-5:]
                reasondescript = 'Description contains: ' + keyword.group()
                portexclusions[host][interface]['reason'].append(reasondescript)
                portexclusions[host][interface]['description']= str(description)

        print("End processing Host - Get Interfaces:", str(host), '\n')

    """
    Export all entries from portexlusions dictionary to portexclusions_ws
    so that port exclusion recommendations show up in the workbook.
    """
    for host, value in portexclusions.items():
        print("Start processing Host - Port Exclusions:", str(host), '\n')
        #print(host)
        for interface in portexclusions[host]:
            #print(host, interface, portexclusions[host][interface]['reason'], portexclusions[host][interface]['description'])
            line = [host, interface, str(portexclusions[host][interface]['reason']), str(portexclusions[host][interface]['description'])]
            portexclusions_ws.append(line)
        print("End processing Host - Port Exclusions:", str(host), '\n')

    """
    Get VLANs... in Napalm-automation:develop train, needed for identifying switchport mode trunk.
    """

    #vlans = accessHosts.run(task=get_vlans, name="Get VLANs")
    #print_result(vlans)

    #portexclusions = mac_results['C9300-48-UXM-1'][0][1]
    #print ('DEBUG EXCLUSIONS>>>> \n', portexclusions)

    #for host, keys in portexclusions.items():
        #print(host, '>', keys)
    #portexclusions[host][iface]['reason'].append('LLDP router Neighbor')

    wb.remove(wb["Sheet"])
    wb.save(wb_name)


""" DELETE
def get_mac_table(task):
    r = task.run(networking.napalm_get, getters=['mac_address_table'], name="Get MAC Table")
    macQ = MacLookup()

    print('-'*10, task.host, '-', task.host.hostname, '- MultiMAC Interfaces','-'*10)
    #transform MAC address table into a new dictionary of interfaces with a list of associated MACs per interface. Only Keep multi mac interfaces
    #task.host['MultiMacports'] = multimacports(r.result)
    #Lookup Vendor OUI from MAC
    vendor_mactable = defaultdict(list)
    interfaces = defaultdict(list)

    #Create dictionary of interfaces and the vendor MACs that are attached.
    for key in r.result['mac_address_table']:
        if not key['interface']:
            continue

        try:
            vendor = macQ.lookup(key['mac'])
        except:
            vendor = "Unknown"

        #print(key['mac'], 'EQUALS > ', vendor, 'ON INTERFACE > ', key['interface'])
        vendor_mactable[key['interface']].append(vendor)

    #From VendorMacTable dictionary, create new list of interfaces that have multiple MACs.
    #Add to Exclusion List
    for iface, value in vendor_mactable.items():
        #print(iface, value)
        if len(value) > 1:
            #print(iface, '>', value)
            interfaces[iface].extend(value)

    print('----- Resolved Vendor MAC Table ----- \n', vendor_mactable)
    print('----- Multi Host Interfaces ----- \n', interfaces)

    task.host['vendormactable'] = vendor_mactable
    task.host['multimacinterfaces'] = interfaces
    #return task.host['exlcusions'][vendor_mactable, interfaces]
    info = {'multimacports': dict(interfaces), 'vendormactable': dict(vendor_mactable)}
    return info
"""
""" DELETE
def get_interfaces(task):
    r = task.run(networking.napalm_get, getters=['interfaces'], name="Get Interfaces")
    # save our values in to the Key 'neighbors'
    task.host['interfaces'] = r.result
    #loop through the neighbors
    print(task.host['interfaces'])
    #print('-'*10, task.host, '-', task.host.hostname, '-'*10)
"""

def get_mac_table_napalm(task):
    task.run(name="Get Mac Table Napalm", task=napalm_get, getters=["mac_address_table"])
    return "Get Mac Table Complete"

def get_facts(task):
    task.run(name="Get facts", task=napalm_get, getters=["facts"])
    return "Get Facts Complete"

def get_lldp_neighbors(task):
    task.run(name="Get LLDP neighbors", task=napalm_get, getters=["lldp_neighbors_detail"])
    return "Get LLDP Neighbors Complete"

def get_interfaces(task):
    task.run(name="Get Interfaces", task=napalm_get, getters=["interfaces"])
    return "get Interfaces Complete"

def get_vlans(task):
    task.run(name="Get VLANs", task=napalm_get, getters=["vlans"])
    return "get VLANs Complete"

"""
Run the main function to pull device information and create the workbook.
"""
create_workbook()
print("Workbook Created")
