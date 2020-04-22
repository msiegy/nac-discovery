"""
    A discovery and collection tool which pulls information from switches and
    provides useful data points and recommendations on which ports
    to apply or exclude 802.1x NAC configurations. Information is exported to an
    Excel workbook.

    Currently Recommendations based on based on LLDP Vendorlookup, RemoteCapability,
    Description keywords, and multiple macs present on port.

    Refactor of collectswitchfacts.py - combined napalm getters and introduced logic for HP Procurve.
    Might require additional logic to run in multivendor scenarios.
"""

from nornir import InitNornir
from nornir.core.exceptions import NornirExecutionError
from nornir.core.exceptions import NornirSubTaskError
from nornir.plugins.tasks import networking
from nornir.plugins.tasks.networking import napalm_get
from nornir.plugins.functions.text import print_result
from mac_vendor_lookup import MacLookup
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font
import re
import time
from nornir.core.deserializer.inventory import InventoryElement

def create_workbook():
    """
    Create an Excel workbook to store values retrieved from switches
    """

    wb = openpyxl.Workbook()
    groupname = "MixGrouping1" #TODO: Replace with function that takes list of location codes.
    wb_name = "MultiDev-NACFACTS -" + groupname + ".xlsx"


    #Create sheets and column headers
    facts_ws = wb.create_sheet("Facts")
    facts_ws.append(['Switch Hostname','Vendor','Model','OS Version','Serial Number','Uptime'])

    interfaces_ws = wb.create_sheet("Interfaces")
    interfaces_ws.append(['Switch', 'Interface name', 'Description', 'Admin Status', 'Oper Status', 'Speed'])

    mactablevendor_ws = wb.create_sheet("Mac Table Vendors")
    mactablevendor_ws.append(['Switch', 'Interface', 'MACaddr', 'Vendor OUI'])

    lldpneighbor_ws = wb.create_sheet("LLDP Neighbors")
    lldpneighbor_ws.append(['Local Switch', 'Local Port', 'Remote System ID', 'Remote System Name', 'Remote System Description', 'Remote Port ID', 'Remote Port Description', 'Remote Capability', 'Remote Vendor'])

    multimacports_ws = wb.create_sheet("Multi Mac Ports")
    multimacports_ws.append(['Switch', 'Interface', 'Count', 'Vendor MACs'])

    portexclusions_ws = wb.create_sheet("Port Exclusion Recommendations")
    portexclusions_ws.append(['Switch', 'Interface', 'Reason', 'Port Description'])

    devicefailures_ws = wb.create_sheet("Failed Devices")
    devicefailures_ws.append(['Switch', 'Hostname', 'Error'])

    """
    Initialize Nornir settings and set the right inventory targets and filters
    """
    nr = InitNornir(config_file="config.yaml", core={"raise_on_error": False})
    #procurve_devices = nr.filter(hostname='10.83.8.163')
    #procurve_devices = nr.filter(site='herndon-dev')
    #procurve_devices = nr.filter(type='network_device')
    #procurve_devices = nr.filter(site='Home')
    procurve_devices = nr.filter(tag='mix')
    #access_devices = nr.filter(type='home')

    #Initialize nested dictionary for tracking recomended ports and reasoning to exclude from NAC.
    portexclusions = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    #portexclusions = {'SwitchName': {'Gi1/1:' {'description': 'trunk to mdf', macvendor: 'VMWare', Reasoning: ['Port is trunk', 'multimacs'] }}

    print('Collecting information from the following Nornir inventory hosts:', procurve_devices.inventory.hosts.keys())

    print("Grabbing Data With Nornir/Napalm - Start Clock\n")
    starttime = time.perf_counter()
    napalm_results = procurve_devices.run(task=napalm_get, getters=['mac_address_table', 'facts', 'lldp_neighbors_detail', 'interfaces'], name="Get Switch info: Facts, MAC Table, LLDP, and Interfaces")
    stoptime = time.perf_counter()
    print(f"Done Grabbing Data\n Execution took: {stoptime - starttime:0.4f} seconds")

    for host, task_results in napalm_results.items():
        #Check for switches that failed and continue. Nornir automatically removes failed devices from future tasks.
        if task_results.failed:
            print(' ######', '\n!Failed Host in task_results:>', host, 'will be removed from future tasks!\n', '######', '\n')
            continue
        print("Start processing Host :", str(host), '\n')
        #print("Task_results are:>\n", task_results)
        facts_result = task_results[0].result['facts']
        #print("Facts_Results are:> (task_results[0].result['facts'] >>\n)", facts_result)
        lldp_result = task_results[0].result['lldp_neighbors_detail']
        #print("lldp_result are:> (task_results[0].result['lldp_neighbors_detail'] >>\n)", lldp_result)
        mactable_result = task_results[0].result['mac_address_table']
        #print("mactable_result are:> (task_results[0].result['mac_address_table'] >>\n)", mactable_result)
        interfaces_result = task_results[0].result['interfaces']
        #print("interfaces_result are:> (task_results[0].result['interfaces'] >>\n)", interfaces_result)

        """PROCESS MAC TABLE RESULTS - Lookup MAC OUI Vendors and Determine Ports with multiple MACs assigned."""
        macQ = MacLookup()
        #loop through each switch in the Nornir Task Result object
        print("Start processing Host - Mac_Results:", str(host), '\n')

        #unnecessary sicne captured above?
        if task_results.failed:
            print(' ######', '\n!Failed Host in task_results (MACTABLE):>', host, 'will be removed from future tasks!\n', '######', '\n')
            continue
        #Store the actual serialized mac table dict for the current switch

        vendor_mactable = defaultdict(list)
        interfaces = defaultdict(list)
        multimacinterfaces = defaultdict(dict)
        #Loop through each Host's MAC Table and Create dictionary of interfaces and the vendor MACs that are attached.
        for entry in mactable_result:
            if not entry['interface']:  #skip mac address not assigned to interfaces
                continue
            try:
                vendor = macQ.lookup(entry['mac'])
            except:
                vendor = "Unknown"

            interface_value = entry['interface']
            vendor_value = vendor
            mac_value = entry['mac']

            #Store relevant values for worksheet row and append to sheet.
            line = [host, interface_value, mac_value, vendor_value]
            print('mactable lookup line:', line)
            mactablevendor_ws.append(line)
            #Append Vendor lookup results to vendor_mactable so we can use them for port exclusion recommendations.
            vendor_mactable[entry['interface']].append(vendor_value)

        #build dictionary of interfaces containing lists of vendors and identify ports with multiple MACs.
        for iface, value in vendor_mactable.items():
            #print(iface, value)
            if len(value) > 1:
                #print(iface, '>', value)
                interfaces[iface].extend(value)

                line = [host, iface, len(interfaces[iface]), str(interfaces[iface])]
                print('multimac line:', line)
                multimacports_ws.append(line)
                #Append to portexlcusions dictionary
                portexclusions[host][iface]['reason'].append('multimac')
        #print('vendor mactable\n\n', vendor_mactable)
        #print('interfact dict \n\n', interfaces)
        print("End Processing Host - Mac_Results: " + str(host) + "\n")

        """
        Get Facts  from all inventory targets using nornir napalm
        task=get_facts and output results to facts_ws
        """


        print("Start processing Host - Get Facts:", str(host), '\n')

        hostname_result = facts_result['hostname']
        vendor_result = facts_result['vendor']
        model_result = facts_result['model']
        version_result = facts_result['os_version']
        serial_result = facts_result['serial_number']
        uptime_result = facts_result['uptime']
        #HP Devices will return a list of interfaces with get_facts
        if facts_result['interface_list']:
            interfacelist_result = facts_result['interface_list']

        line = [host, vendor_result, model_result, version_result, serial_result, uptime_result]
        print('Facts line:', line)
        print(interfacelist_result)

        facts_ws.append(line)
        print("End Processing Host - Get Facts: " + str(host) + "\n")


        """PROCESS LLDP NEIGHBOR DETAIL RESULTS - ."""

        print("Start processing Host - Get LLDP Neighbors:", str(host), '\n')
        for interface in lldp_result:
            #print(lldp_detail)

            remotesysid = lldp_result[interface][0]['remote_chassis_id']
            remotesysname = lldp_result[interface][0]['remote_system_name']
            remotesysdescription = lldp_result[interface][0]['remote_system_description']
            remoteportid = lldp_result[interface][0]['remote_port']
            remoteportdesc = lldp_result[interface][0]['remote_port_description']
            remotecapability = lldp_result[interface][0]['remote_system_capab']
            try:
                remotevendor = macQ.lookup(remotesystemid)
            except:
                remotevendor = "Unknown"

            line = [host, interface, remotesysid, remotesysname, remotesysdescription, remoteportid, remoteportdesc, str(remotecapability), remotevendor]
            print('lldp neighbor line::', line)
            lldpneighbor_ws.append(line)

            if ('router' in remotecapability) or ('bridge' in remotecapability):
                #TODO: generalize for all interfaces and move to function
                if re.search('TenGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Te' + digits.group()
                elif re.search('TwoGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Tw' + digits.group()
                elif re.search('^Gigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Gi' + digits.group()

                portexclusions[host][interface]['reason'].append('LLDP Neighbor' + str(remotecapability))
                #print(host, interface, remotecapability)

        print("End Processing Host - Get LLDP Neighors: " + str(host) + "\n")


        """
        Get Interfaces, check descriptions for keywords and append to port exclusions.
        """
        print("Start processing Host - Get Interfaces:", str(host), '\n')

        for interface in interfaces_result:
            #if 'C9500-16X' in host:

            interface_id = interface
            adminstatus = interfaces_result[interface]['is_enabled']
            operstatus = interfaces_result[interface]['is_up']
            description = interfaces_result[interface]['description']
            speed = interfaces_result[interface]['speed']

            line = [host, interface_id, description, adminstatus, operstatus, speed]
            print('Interfaces line:', line)

            interfaces_ws.append(line)

            #Check for Exclusion keywords and add interfaces to portexclusion dictionary then append to portexlusion_ws.
            #TODO: Replace search literals with variable at top for quicker modification.
            keyword = re.search('(ASR|ENCS|UPLINK|CIRCUIT|ISP|SWITCH|TRUNK|ESXI|VMWARE)', str(description), re.IGNORECASE)
            if keyword:
                #Normalize Interface names because different napalm getters return full interfaces name and some return shortened names which result in multiple dictionary keys being created.
                #TODO: generalize for all interfaces and move to function
                if re.search('TenGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Te' + digits.group()
                elif re.search('TwoGigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Tw' + digits.group()
                elif re.search('^Gigabit', str(interface), re.IGNORECASE):
                    digits = re.search('([0-9]*\/?[0-9]*\/?[0-9]*$)', str(interface))
                    interface = 'Gi' + digits.group()

                reasondescript = 'Description contains: ' + keyword.group()
                portexclusions[host][interface]['reason'].append(reasondescript)
                portexclusions[host][interface]['description']= str(description)

        print("End processing Host - Get Interfaces:", str(host), '\n')


        """
        Export all entries from portexlusions dictionary to portexclusions_ws
        so that port exclusion recommendations show up in the workbook.
        """
        for host, value in portexclusions.items():
            print("Start processing Host - Port Exclusions:", str(host), '\n')
            #print(host)
            for interface in portexclusions[host]:
                line = [host, interface, str(portexclusions[host][interface]['reason']), str(portexclusions[host][interface]['description'])]
                print("Port Exclusions line:", line)
                portexclusions_ws.append(line)
            print("End processing Host - Port Exclusions:", str(host), '\n')

    #check if there are any failed hosts and save failed switches to worksheet
    if nr.data.failed_hosts:
        print("The following switches failed during a task and were not added to the workbook:", nr.data.failed_hosts)
        for host in nr.data.failed_hosts:
            error = napalm_results[host][0].exception
            #print(procurve_devices.inventory.hosts[host].keys())
            hostname = procurve_devices.inventory.get_hosts_dict()[host]['hostname'])
            line = [host, hostname, str(error)]
            devicefailures_ws.append(line)

    wb.remove(wb["Sheet"])
    #catch potential save errors on workbook.
    try:
        wb.save(wb_name)
        print("\nWorkbook Created")
    except Exception as e:
        print("\n######", e , "######\nFailed to Save workbook, please close it if open and ensure you have access to save location")
"""
Run the main function to pull device information and create the workbook.
"""
create_workbook()
